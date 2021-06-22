from django.shortcuts import render, redirect
from eval.Utils import getuname
from eval_accounts.models import AccountDetails, EvaluationPeriod
from eval_main.models import MainHrHrMaster
from .models import EmployeeEval, EvaluationCriteria, Evaluation, Comments
from eval_app.forms import EmployeeEvalForm, EvaluationForm, CommentsForm
from django.forms import inlineformset_factory, modelformset_factory
from django.http import HttpResponse
from django.contrib.staticfiles.storage import staticfiles_storage
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill
# Create your views here.


def error(request):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Error Page'})
    
    return render(request,'eval_app/error.html', context)


def index(request):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Dashboard'})
    return render(request,'eval_app/index.html', context )


def MyEvaluation(request):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'My Evaluation       ( ' + evaluation_period.period+', '+ str(evaluation_period.yr) + ' )'})
    
    if len(MainHrHrMaster.objects.using('main').filter(mhi_id=uname).exclude(employee_status='Admin')) == 0:
        return render(request,'eval_app\error.html',context)

    umaster = MainHrHrMaster.objects.using('main').get(mhi_id=uname)
    context.update({'umaster' : umaster})
    
    if len(EmployeeEval.objects.filter(employee_id=umaster.idno).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period)) !=0:
        EmployeeEvalinstance = EmployeeEval.objects.get(employee_id=umaster.idno,yr=evaluation_period.yr, period=evaluation_period.period)
        if udetails.access.id == 5 or udetails.access.id == 4:
            if EmployeeEvalinstance.user_submit =='Submitted':
                context.update({'reason' : 'Evaluation was already submitted' })
                return render(request, 'eval_app/info.html', context)
        if EmployeeEvalinstance.lead_submit == 'Submitted':
            context.update({'reason' : 'Evaluation was already submitted' })
            return render(request, 'eval_app/info.html', context)
    else:
        b = EmployeeEval(employee_id=umaster.idno, yr=evaluation_period.yr,period=evaluation_period.period)
        b.save()

    EmployeeEvalinstance = EmployeeEval.objects.get(employee=umaster.idno,yr=evaluation_period.yr, period=evaluation_period.period)
    form = EmployeeEvalForm(instance=EmployeeEvalinstance)
    if request.method =='POST':
        form =EmployeeEvalForm(request.POST or None, instance=EmployeeEvalinstance)
        if form.is_valid():
            form.save()   
            return redirect('eval-MyEvaluation')
        else:
            print(form.errors)

    context.update({'form' : form})
    return render(request,'eval_app/employee_MyEvaluation.html', context)


def EvaluationList(request):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Evaluation       ( ' + evaluation_period.period+', '+ str(evaluation_period.yr) + ' )'})
    if len(MainHrHrMaster.objects.using('main').filter(mhi_id=uname)) != 0:
        umaster = MainHrHrMaster.objects.using('main').get(mhi_id=uname)
        context.update({'umaster' : umaster})

    if udetails.access.id  == 5 or udetails.access.id == 4:
        context.update({'error' : 'You are not allowed to access this site'})
        return render(request,'eval_app/error.html',context)
    
    elist=[]
    if udetails.access.id == 2 :
        evaluation_list = AccountDetails.objects.exclude(access_id=7).exclude(access_id=1).exclude(access_id=2).exclude(employee_id=udetails.employee_id)
    else:
        evaluation_list = AccountDetails.objects.filter(project=udetails.project).exclude(access_id=7).exclude(access_id=1).exclude(access_id=2).exclude(employee_id=udetails.employee_id)
    
    for i in evaluation_list:
        if len(EmployeeEval.objects.filter(employee_id=i.employee_id).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period)) == 0:
            user_status = 'On-Going'
            lead_status = 'On-Going'
        else:
            user_status = EmployeeEval.objects.get(employee_id=i.employee_id,yr=evaluation_period.yr,period=evaluation_period.period).user_submit
            lead_status = EmployeeEval.objects.get(employee_id=i.employee_id,yr=evaluation_period.yr,period=evaluation_period.period).lead_submit

        info ={
            "employee_id": i.employee_id,
            "last_name" : i.last_name,
            "first_name" : i.first_name,
            "nick_name" : i.nick_name,
            "project" : i.project,
            "area" : i.area,
            "user_status" : user_status,
            "lead_status" : lead_status
        }
        elist.append(info)
    context.update({'elist': elist})
    
    return render(request,'eval_app/evaluation.html', context)


def EvaluationLead(request, idno):
    context = {}
    idno = idno
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Evaluation       ( ' + evaluation_period.period+', '+ str(evaluation_period.yr) + ' )'})
    eval_udetails = AccountDetails.objects.get(employee_id=idno)
    context.update({'eval_udetails' : eval_udetails})
    
    if udetails.access.id  == 5 or udetails.access.id == 4:
        context.update({'error' : 'You are not allowed to access this site'})
        return render(request,'eval_app/error.html',context)
    
    if udetails.access.id == 3:
        
        
        if udetails.project != eval_udetails.project :
            context.update({'error' : 'You are not allowed to access this employee'})
            return render(request,'eval_app/error.html',context)
    
    if len(EmployeeEval.objects.filter(employee_id=idno).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period)) !=0:
        EmployeeEvalinstance = EmployeeEval.objects.get(employee_id=idno,yr=evaluation_period.yr, period=evaluation_period.period)
        if udetails.access.id == 5:
            if EmployeeEvalinstance.user_submit =='Submitted':
                context.update({'reason' : 'Evaluation was already submitted' })
                return render(request, 'eval_app/info.html', context)
        if udetails.access.id !=7 and udetails.access.id !=2:
            if EmployeeEvalinstance.lead_submit == 'Submitted':
                context.update({'reason' : 'Evaluation was already submitted' })
                return render(request, 'eval_app/info.html', context)
    else:
        b = EmployeeEval(employee_id=idno, yr=evaluation_period.yr,period=evaluation_period.period)
        b.save()
    
    EmployeeEvalinstance = EmployeeEval.objects.get(employee=idno,yr=evaluation_period.yr, period=evaluation_period.period)
    form = EmployeeEvalForm(instance=EmployeeEvalinstance)
    if request.method =='POST':
        form =EmployeeEvalForm(request.POST or None, instance=EmployeeEvalinstance)
        if form.is_valid():
            form.save()   
            return redirect('eval-EvaluationLead', idno)
        else:
            print(form.errors)

    context.update({'form' : form})
    return render(request, 'eval_app/evaluation_lead.html', context)


def MySkills(request):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'My Skills       ( ' + evaluation_period.period+', '+ str(evaluation_period.yr) + ' )'})

    if len(MainHrHrMaster.objects.using('main').filter(mhi_id=uname)) == 0:
        context.update({'error' : 'You are not allowed'})
        return redirect('eval-error',context)

    umaster = MainHrHrMaster.objects.using('main').get(mhi_id=uname)
    context.update({'umaster' : umaster})

    criterias = EvaluationCriteria.objects.filter(group__contains=udetails.skill_set)
    
    if len(AccountDetails.objects.filter(project=udetails.project).filter(area=udetails.area).filter(access_id=4)) == 0:
        arealead =''
    else:
        arealead = AccountDetails.objects.get(project=udetails.project,area=udetails.area, access_id=4).employee_id

    if len(AccountDetails.objects.filter(project=udetails.project).filter(access_id=3).filter(area=udetails.area)) == 0:
        projectlead =''
    else:
        projectlead = AccountDetails.objects.get(project=udetails.project, access_id=3,area=udetails.area).employee_id
    
    if len(AccountDetails.objects.filter(access_id=2)) == 0:
        manager =''
    else:
        manager = AccountDetails.objects.get(access_id=2).employee_id
    
    EvaluationFormSet = modelformset_factory(Evaluation, form=EvaluationForm, extra=0)
    for i in criterias:        
        if len(Evaluation.objects.filter(employee_id=udetails.employee_id).filter(criteria=i).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period)) == 0:
            b = Evaluation(employee_id=udetails.employee_id,criteria=i,self_rating='#',area_lead_id=arealead, area_lead_rating='#', project_lead_id=projectlead, 
            project_lead_rating='#', manager_id=manager, manager_rating='#', actions='', yr=evaluation_period.yr, period=evaluation_period.period)
            b.save()
    
    if request.method == 'POST':
        formset = EvaluationFormSet(request.POST, queryset=Evaluation.objects.filter(employee_id=udetails.employee_id).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period))
        
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.employee_id = udetails.employee_id 
                instance.save()
        else:
            print(formset.errors)
      
        if udetails.access.id ==3 :
            b = Evaluation.objects.filter(employee_id=udetails.employee_id).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period)
            for i in b:
                i.project_lead_rating = i.self_rating
                i.save()
        if udetails.access.id ==4 :
            b = Evaluation.objects.filter(employee_id=udetails.employee_id).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period)
            for i in b:
                i.area_lead_rating = i.self_rating
                i.save()

     
    formset = EvaluationFormSet(queryset=Evaluation.objects.filter(employee_id=udetails.employee_id).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period))

    context.update({'formset' : formset})

    return render(request,'eval_app/employee_MySkills.html', context)


def SkillList(request):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Skills       ( ' + evaluation_period.period+', '+ str(evaluation_period.yr) + ' )'})
    if len(MainHrHrMaster.objects.using('main').filter(mhi_id=uname)) != 0:
        umaster = MainHrHrMaster.objects.using('main').get(mhi_id=uname)
        context.update({'umaster' : umaster})

    if udetails.access.id  == 5 :
        context.update({'error' : 'You are not allowed to access this site'})
        return render(request,'eval_app/error.html',context)

    if udetails.access.id == 4 :
        evaluation_list= AccountDetails.objects.filter(area=udetails.area).exclude(employee_id=udetails.employee_id)
    elif udetails.access.id == 3 :
        evaluation_list = AccountDetails.objects.filter(project=udetails.project).exclude(employee_id=udetails.employee_id).exclude(access_id=2)
    elif udetails.access.id == 2 :
        evaluation_list = AccountDetails.objects.exclude(access_id=1).exclude(access_id=2).exclude(access_id=6).exclude(access_id=7)
    elif udetails.access.id == 7:
        evaluation_list = AccountDetails.objects.filter(project=udetails.project).exclude(access_id=1).exclude(access_id=2).exclude(access_id=6).exclude(access_id=7)
    elist=[]
    for i in evaluation_list:
        info ={
            "employee_id": i.employee_id,
            "last_name" : i.last_name,
            "first_name" : i.first_name,
            "nick_name" : i.nick_name,
            "project" : i.project,
            "area" : i.area,
        }
        elist.append(info)
    context.update({'elist': elist})


    return render(request,'eval_app/skills.html', context)


def SkillLead(request,idno):
    context = {}
    idno = idno
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Skills       ( ' + evaluation_period.period+', '+ str(evaluation_period.yr) + ' )'})
    eval_udetails = AccountDetails.objects.get(employee_id=idno)
    context.update({'eval_udetails' : eval_udetails})

    if udetails.access.id  == 5 :
        context.update({'error' : 'You are not allowed to access this site'})
        return render(request,'eval_app/error.html',context)
    
    if udetails.access.id == 4:
        
        if udetails.area != eval_udetails.area :
            context.update({'error' : 'You are not allowed to access this employee'})
            return render(request,'eval_app/error.html',context)
    
    if udetails.access.id == 3:
        if udetails.project != eval_udetails.project :
            context.update({'error' : 'You are not allowed to access this employee'})
            return render(request,'eval_app/error.html',context)

    
    criterias = EvaluationCriteria.objects.filter(group__contains=eval_udetails.skill_set)
    if len(AccountDetails.objects.filter(project=eval_udetails.project).filter(area=eval_udetails.area).filter(access_id=4)) == 0:
        arealead =''
        arealeadname = ''
    else:
        arealead = AccountDetails.objects.get(project=eval_udetails.project,area=eval_udetails.area, access_id=4).employee_id
        arealeadname = AccountDetails.objects.get(project=eval_udetails.project,area=eval_udetails.area, access_id=4).nick_name
    context.update({'arealeadname' : arealeadname})
    if len(AccountDetails.objects.filter(project=eval_udetails.project).filter(access_id=3)) == 0:
        projectlead =''
        projectleadname =AccountDetails.objects.get(project=eval_udetails.project,access_id=3 ).nick_name
    else:
        projectlead = AccountDetails.objects.get(project=eval_udetails.project, access_id=3).employee_id
        projectleadname = AccountDetails.objects.get(project=eval_udetails.project, access_id=3).nick_name
    context.update({'projectleadname' : projectleadname})
    if len(AccountDetails.objects.filter(access_id=2)) == 0:
        manager =''
        managername =''
    else:
        manager = AccountDetails.objects.get(access_id=2).employee_id
        managername = AccountDetails.objects.get(access_id=2).nick_name
    context.update({'managername' : managername})
    EvaluationFormSet = modelformset_factory(Evaluation, form=EvaluationForm, extra=0)
    for i in criterias:        
        if len(Evaluation.objects.filter(employee_id=eval_udetails.employee_id).filter(criteria=i).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period)) == 0:
            b = Evaluation(employee_id=eval_udetails.employee_id,criteria=i,self_rating='#',area_lead_id=arealead, area_lead_rating='#', project_lead_id=projectlead, 
            project_lead_rating='#', manager_id=manager, manager_rating='#', actions='', yr=evaluation_period.yr, period=evaluation_period.period)
            b.save()


    if request.method == 'POST':
        formset = EvaluationFormSet(request.POST, queryset=Evaluation.objects.filter(employee_id=eval_udetails.employee_id).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period))
        
        if formset.is_valid():
            instances = formset.save(commit=False)
            for instance in instances:
                instance.employee_id = eval_udetails.employee_id 
                instance.save()
        else:
            print(formset.errors)    

    formset = EvaluationFormSet(queryset=Evaluation.objects.filter(employee_id=eval_udetails.employee_id).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period))

    context.update({'formset' : formset})

    return render(request,'eval_app/skills_lead.html',context)


def MyComments(request):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'My Comments       ( '+ str(evaluation_period.yr) + ' )'})


    
    
    if len(MainHrHrMaster.objects.using('main').filter(mhi_id=uname).exclude(employee_status='Admin')) == 0:
        return render(request,'eval_app\error.html',context)

    umaster = MainHrHrMaster.objects.using('main').get(mhi_id=uname)
    context.update({'umaster' : umaster})

    if len(Comments.objects.filter(employee_id=umaster.idno).filter(yr=evaluation_period.yr)) !=0:
        EmployeeCommentinstance = Comments.objects.get(employee_id=umaster.idno,yr=evaluation_period.yr)
    else:
        b = Comments(employee_id=umaster.idno, yr=evaluation_period.yr)
        b.save()
    
    EmployeeCommentinstance = Comments.objects.get(employee_id=umaster.idno,yr=evaluation_period.yr)
    
    print(EmployeeCommentinstance.status)

    if EmployeeCommentinstance.status == 'Submitted':
        context.update({'reason' : 'Evaluation was already submitted' })
        return render(request, 'eval_app/info.html', context)



    context.update({'EmployeeCommentinstance' : EmployeeCommentinstance})
    form = CommentsForm(instance=EmployeeCommentinstance)
    if request.method =='POST':
        form =CommentsForm(request.POST or None, instance=EmployeeCommentinstance)
        if form.is_valid():
            form.save()   
            return redirect('eval-MyComments')
        else:
            print(form.errors)

    context.update({'form' : form})

    return render(request,'eval_app/employee_MyComments.html', context)


def CommentsList(request):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Comments       ( '+ str(evaluation_period.yr) + ' )'})
    if len(MainHrHrMaster.objects.using('main').filter(mhi_id=uname)) != 0:
        umaster = MainHrHrMaster.objects.using('main').get(mhi_id=uname)
        context.update({'umaster' : umaster})

    if udetails.access.id  != 2 :
        context.update({'error' : 'You are not allowed to access this site'})
        return render(request,'eval_app/error.html',context)

    evaluation_list = AccountDetails.objects.exclude(access_id=1).exclude(access_id=2).exclude(access_id=6).exclude(access_id=7).order_by('nick_name')
    
    elist=[]
    for i in evaluation_list:
        if len(Comments.objects.filter(employee_id=i.employee_id).filter(yr=evaluation_period.yr).filter(status='Submitted')) !=0:
            color = "green"
        else:
            color = "red"
        info ={
            "employee_id": i.employee_id,
            "last_name" : i.last_name,
            "first_name" : i.first_name,
            "nick_name" : i.nick_name,
            "project" : i.project,
            "area" : i.area,
            "pic" : i.pic,
            "color" : color,
        }
        elist.append(info)
    context.update({'elist': elist})
    return render(request,'eval_app/comments.html', context)
    

def CommentLead(request, idno):
    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    evaluation_period = EvaluationPeriod.objects.get()
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Comments       ( '+ str(evaluation_period.yr) + ' )'})
    eval_udetails = AccountDetails.objects.get(employee_id=idno)
    context.update({'eval_udetails' : eval_udetails})
    
    if udetails.access.id  != 2 :
        context.update({'error' : 'You are not allowed to access this site'})
        return render(request,'eval_app/error.html',context)
    
    if len(Comments.objects.filter(employee_id=idno).filter(yr=evaluation_period.yr).filter(status='Submitted')) == 0:
        context.update({'error' : 'Employee has not submitted his/her comments'})
        return render(request,'eval_app/error.html',context)


    EmployeeCommentinstance = Comments.objects.get(employee_id=idno,yr=evaluation_period.yr)
    form = CommentsForm(instance=EmployeeCommentinstance)
    context.update({'form' : form})
    return render(request,'eval_app/comments_lead.html', context)


def AccessDenied(request):
    return render(request,'eval_app/access_denied.html')


def HrSync(request):

    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    context.update({'udetails' : udetails})
    context.update({'loc' : 'Sync'})

    if udetails.admin == 0:
        context.update({'error' : 'You are not allowed'})
        return render(request,'eval_app/error.html', context)

    a = MainHrHrMaster.objects.using('main').all()

    for i in a: 
        emplstat = str(i.employee_status)
        
        if emplstat == 'Resigned' or emplstat == 'Terminated' or emplstat == 'Outsourced':
            if len(AccountDetails.objects.filter(mhi_id=i.mhi_id)) !=0:
                d = AccountDetails.objects.get(mhi_id=i.mhi_id)
                d.delete()
        else:
            if len(AccountDetails.objects.filter(mhi_id=i.mhi_id)) == 0:
                b = AccountDetails(mhi_id=i.mhi_id, last_name= i.last_name, first_name=i.first_name,
                    nick_name=i.nick_name, pic=i.picture, admin=0, employee_id=i.idno, project=i.project.project_code_description )
                b.save()
            else:
                d = AccountDetails.objects.get(mhi_id=i.mhi_id)
                #d.project = i.project.project_code_description
                d.employee_id = i.idno
                d.first_name = i.first_name
                d.last_name = i.last_name
                d.nick_name = i.nick_name
                d.save()
                
    return render(request,'eval_app/sync.html', context)





def WriteToExcel(request, idno):

    context = {}
    uname = getuname(request)
    udetails = AccountDetails.objects.get(mhi_id=uname)
    evaluation_period = EvaluationPeriod.objects.get()
    eval_udetails = AccountDetails.objects.get(employee_id=idno)
    
    if len(MainHrHrMaster.objects.using('main').filter(mhi_id=eval_udetails.mhi_id)) != 0:
        umaster = MainHrHrMaster.objects.using('main').get(mhi_id=eval_udetails.mhi_id)

    
    if udetails.access.id  != 2 :
        context.update({'error' : 'You are not allowed to access this site'})
        return render(request,'eval_app/error.html',context)
    
    
    if eval_udetails.skill_set == 'I.T.' :
        fname = staticfiles_storage.path('eval_app/xlsx/IT.xlsx')
    elif eval_udetails.skill_set == 'ADMIN':    
        fname = staticfiles_storage.path('eval_app/xlsx/ADMIN.xlsx')

    skills = Evaluation.objects.filter(employee_id=eval_udetails.employee_id).filter(yr=evaluation_period.yr).filter(period=evaluation_period.period)
    
    
    
    wb = load_workbook(filename=fname)
    ws = wb.active
    ws['D2'] = eval_udetails.last_name+', '+eval_udetails.first_name+' ('+eval_udetails.nick_name+')'
    ws['D3'] = idno
    
    
    rw = 11
    for i in skills:
        if i.self_rating == '1':
            ws['G'+str(rw)].fill = PatternFill(start_color="D0CDCD", end_color="D0CDCD", fill_type = "solid")
        elif i.self_rating == '2':
            ws['H'+str(rw)].fill = PatternFill(start_color="D0CDCD", end_color="D0CDCD", fill_type = "solid")
        elif i.self_rating == '3':
            ws['I'+str(rw)].fill = PatternFill(start_color="D0CDCD", end_color="D0CDCD", fill_type = "solid")
        elif i.self_rating == '4':
            ws['J'+str(rw)].fill = PatternFill(start_color="D0CDCD", end_color="D0CDCD", fill_type = "solid")
        else:
            ws['E'+str(rw)].fill = PatternFill(start_color="D0CDCD", end_color="D0CDCD", fill_type = "solid")
        
        if i.area_lead_rating == '1':
            ws['M'+str(rw)].fill = PatternFill(start_color="0b3b72", end_color="0b3b72", fill_type = "solid")
        elif i.area_lead_rating == '2':
            ws['N'+str(rw)].fill = PatternFill(start_color="0b3b72", end_color="0b3b72", fill_type = "solid")
        elif i.area_lead_rating == '3':
            ws['O'+str(rw)].fill = PatternFill(start_color="0b3b72", end_color="0b3b72", fill_type = "solid")
        elif i.area_lead_rating == '4':
            ws['P'+str(rw)].fill = PatternFill(start_color="0b3b72", end_color="0b3b72", fill_type = "solid")
        else:
            ws['K'+str(rw)].fill = PatternFill(start_color="0b3b72", end_color="0b3b72", fill_type = "solid")

        if i.project_lead_rating == '1':
            ws['S'+str(rw)].fill = PatternFill(start_color="d48d4f", end_color="d48d4f", fill_type = "solid")
        elif i.project_lead_rating == '2':
            ws['T'+str(rw)].fill = PatternFill(start_color="d48d4f", end_color="d48d4f", fill_type = "solid")
        elif i.project_lead_rating == '3':
            ws['U'+str(rw)].fill = PatternFill(start_color="d48d4f", end_color="d48d4f", fill_type = "solid")
        elif i.project_lead_rating == '4':
            ws['V'+str(rw)].fill = PatternFill(start_color="d48d4f", end_color="d48d4f", fill_type = "solid")
        else:
            ws['Q'+str(rw)].fill = PatternFill(start_color="d48d4f", end_color="d48d4f", fill_type = "solid")

        if i.manager_rating == '1':
            ws['Y'+str(rw)].fill = PatternFill(start_color="0ec966", end_color="0ec966", fill_type = "solid")
        elif i.manager_rating == '2':
            ws['Z'+str(rw)].fill = PatternFill(start_color="0ec966", end_color="0ec966", fill_type = "solid")
        elif i.manager_rating == '3':
            ws['AA'+str(rw)].fill = PatternFill(start_color="0ec966", end_color="0ec966", fill_type = "solid")
        elif i.manager_rating == '4':
            ws['AB'+str(rw)].fill = PatternFill(start_color="0ec966", end_color="0ec966", fill_type = "solid")
        else:
            ws['W'+str(rw)].fill = PatternFill(start_color="0ec966", end_color="0ec966", fill_type = "solid")
      
        rw=rw+1

    response = HttpResponse(content=save_virtual_workbook(wb), content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename=' + idno + '.xlsx'
    return response