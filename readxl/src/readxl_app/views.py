from django.shortcuts import render
from openpyxl import load_workbook
# Create your views here.

def index(request):
    workbook = load_workbook("C:\Bin\Book1.xlsx")
    

    sheet = workbook.active
    
    # print(sheet.cell(row=1,column=1).value)
    # print(sheet.cell(row=2,column=1).value)

    for row in sheet.iter_rows(min_row=1,max_row=10,min_col=1,max_col=2,values_only=True):
        print(str(row[0])+' - ' +str(row[1]))    


    return render(request, 'readxl_app/index.html')
